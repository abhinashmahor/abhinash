# -*- coding: utf-8 -*-

import requests
import xlrd
import time
import openpyxl
from urllib.parse import urlencode
from openpyxl.styles import Font
API_KEY = "AIzaSyCtbt_IRW-_Hj2EizWFyL18puPUZ06c8_c"

class Search_functions:
    
    
    
    
    def make_a_url(search_text, centre_lat,centre_lng, radius):
        basic_url_str = f'https://maps.googleapis.com/maps/api/place/nearbysearch/json'
        params_2 = {'key' : API_KEY,
                    "radius": radius ,
                    "keyword": search_text
                    }
        print(params_2)
        str1 = f"{centre_lat},{centre_lng}"
        params_2['location'] =str1
        params_encoded = urlencode(params_2)
        params_url = f"{basic_url_str}?{params_encoded}"
        print(params_url)
        return(params_url)   
        
        
    def write_in_excel(result_list,const, search, city, sheet2):
        #print(result_list)
        for i in range(len(result_list)):
            print(i)
            #my_workbook = openpyxl.load_workbook('KFC_pune.xlsx')
            #sheet = my_workbook.active
            max_roow = sheet2.max_row
            try:
                sheet2.cell(row = max_roow+1, column=1).value = result_list[i]['name']
            except:
                pass
            try:
                sheet2.cell(row=max_roow+1,column=2).value = result_list[i]['plus_code']['compound_code']
            except:
                try:
                    sheet2.cell(row=max_roow+1,column=2).value = result_list[i]['plus_code']['global_code']
                except:
                    pass
            try:
                sheet2.cell(row=max_roow+1,column=3).value = result_list[i]["rating"]
            except:
                pass
            try:
                sheet2.cell(row=max_roow+1,column=4).value = result_list[i]['place_id']
            except:
                pass
            base_str = 'https://www.google.com/maps/place/?q=place_id:'
            f_url_place_id = f"{base_str}{result_list[i]['place_id']}"
            try:
                sheet2.cell(row=max_roow+1,column=5).value = f_url_place_id
            except:
                pass
            
            try:
                sheet2.cell(row=max_roow+1,column=6).value = result_list[i]['business_status']
            except:
                pass
            try:
                sheet2.cell(row=max_roow+1,column=7).value = result_list[i]['user_ratings_total']
            except:
                pass
            try:
                sheet2.cell(row=max_roow+1,column=8).value = result_list[i]['vicinity']
            except:
                pass
            print(result_list[i]['name'])
            try:
                sheet2.cell(row=max_roow+1,column=9).value = result_list[i]['geometry']['location']['lat']
                sheet2.cell(row=max_roow+1,column=10).value = result_list[i]['geometry']['location']['lng']            
            except:
                pass
            print(result_list[i]['name'])
            sheet2.cell(row=max_roow+1,column=12).value = search
            sheet2.cell(row=max_roow+1,column=13).value=city
            
            if(const ==1):
                sheet2.cell(row=max_roow+1,column=18).value = 'next_page_results'
            if(const ==2):
                sheet2.cell(row=max_roow+1,column=18).value = 'next_page_results 3'
            #my_workbook.save('KFC_pune.xlsx')
            
    def read_cord_slow(city_name, search_items, s2):
        print('in read')
        wb = xlrd.open_workbook('grid.xlsx')
        sheet = wb.sheet_by_index(0)
     
        # For row 0 and column 0
        #print(sheet.cell_value(30, 0))
        print(sheet.nrows)
        
        print(sheet.ncols)
        for search_type in search_items:
            for i in range(2,sheet.nrows):
                print(i)
                latx = sheet.cell_value(i,0)
                lngx = sheet.cell_value(i,1)
                url = Search_functions.make_a_url(search_type,latx, lngx, 800)
                print(url)
                r = requests.get(url)
                if(r.json()['results']!=[]):
                    for key in r.json()['results'][0]:
                        print(key)
                    
                    print(url)
                    print(r.json())
                    Search_functions.write_in_excel(r.json()['results'],0,search_type, city_name,s2)
                    if('next_page_token' in r.json().keys()):
                        time.sleep(2)
                        url_base = 'https://maps.googleapis.com/maps/api/place/nearbysearch/json?pagetoken='+r.json()['next_page_token']+'&key='+API_KEY
                        r2 = requests.get(url_base)
                        print(url_base)
                        if(r2.json()['results']!=[]):
                            Search_functions.write_in_excel(r2.json()['results'],1, search_type, city_name,s2)
                        if('next_page_token' in r2.json().keys()):
                            time.sleep(2)
                            url_base = 'https://maps.googleapis.com/maps/api/place/nearbysearch/json?pagetoken='+r2.json()['next_page_token']+'&key='+API_KEY
                            r3 = requests.get(url_base)
                            print(url_base)
                            if(r3.json()['results']!=[]):
                                Search_functions.write_in_excel(r3.json()['results'],2, search_type, city_name, s2)
    
    
    
    
    def write_headers(name):
        wb= openpyxl.load_workbook(name)
        sheet2 = wb.active
        sheet2.cell(row=1,column=1).value = 'Name'
        sheet2.cell(row=1,column=1).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=2).value = 'Plus Code'
        sheet2.cell(row=1,column=2).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=3).value = 'Rating'
        sheet2.cell(row=1,column=3).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=4).value = 'Place ID'
        sheet2.cell(row=1,column=4).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=5).value = 'Link'
        sheet2.cell(row=1,column=5).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=6).value = 'Business Status'
        sheet2.cell(row=1,column=6).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=7).value = 'Number of reviews'
        sheet2.cell(row=1,column=7).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=8).value = 'Vicinity'
        sheet2.cell(row=1,column=8).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=9).value = 'Lattitude'
        sheet2.cell(row=1,column=9).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=10).value = 'Longitude'
        sheet2.cell(row=1,column=10).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=11).value = 'Geocode Link'
        sheet2.cell(row=1,column=11).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=12).value = 'RG address'
        sheet2.cell(row=1,column=12).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=13).value = 'Pin Code'
        sheet2.cell(row=1,column=13).font = Font(bold =True, underline = 'single')
        sheet2.cell(row=1,column=14).value = 'Short Address'
        sheet2.cell(row=1,column=14).font = Font(bold =True, underline = 'single')  
        wb.save(name)
