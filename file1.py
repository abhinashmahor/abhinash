# -*- coding: utf-8 -*-
"""
Created on Tue Nov  9 14:29:24 2021

@author: Lenovo
"""
import requests
import xlrd
import time
import openpyxl
from urllib.parse import urlencode
from openpyxl.styles import Font
import openpyxl




class Make_grid:
    '''
    Takes coordinates of three vertices as input
    takes horizontal distance and vertical distance as input
         
    
    
         d1   
    P1_________P2
    |          |
    |          | d2
    |          |
    |__________P3
               
    
    
    
    Returns a collection of points
    
    '''
    
    def __init__(self,p1,p2,p3, h_dist, v_dist):
        self.p1 = p1
        self.p2 = p2
        self.p3 = p3
        self.h_dist = h_dist
        self.v_dist = v_dist
    
    def method1(self):
        lat1 = self.p1[0]
        lng1 = self.p1[1]
        lat2 = self.p3[0]
        lng2 = self.p2[1]
        wb = openpyxl.load_workbook('grid.xlsx')
        ws = wb.active
        for i in range(2,ws.max_row+1):
            ws.cell(i,1).value = None
            ws.cell(i,2).value = None
        row = 2
        #print(lat1,lat2,lng1,lng2)
        temp_lat=lat2
        while(temp_lat<lat1):
            temp_lng = lng1
            while(temp_lng<lng2):
                #print(temp_lat)
                #print(temp_lng)
                temp_lng+=(lng2-lng1)/self.h_dist
                #print('\n')
                string =f"['A','B', {temp_lat}, {temp_lng}],"
                ws.cell(row, 1).value = temp_lat
                ws.cell(row,2).value = temp_lng
                
                print(string)
                row+=1
            temp_lat+=(lat1-lat2)/self.v_dist
        wb.save('grid.xlsx')
